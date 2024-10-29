from io import BytesIO
import os
from django.db import IntegrityError
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from openpyxl import load_workbook
from tablib import Dataset
from dateutil import parser as date_parser
import tablib
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from accounts.models import Teacher
from students.models import Student
from classrooms.models import Classroom
import datetime
from django.db.models import Q

from django.contrib.auth.models import Group
from classrooms.models import Classroom
from django.core.exceptions import PermissionDenied


def download_template(request):
    # Logic to serve the download link for the template file
    existing_wb = load_workbook(
        'SIMS/static/media/VRCNHS_STUDENT_TEMPLATE(CLEAR).xlsx')

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    existing_wb.save(buffer)
    buffer.seek(0)

    response = FileResponse(buffer, as_attachment=True,
                            filename='Students_Template.xlsx')

    return response


def import_students(request):
    if request.method == 'POST':
        dataset = Dataset()
        new_student_file = request.FILES['myfile']

        # Check if the uploaded file is an Excel file
        if not new_student_file.name.endswith('xlsx'):
            messages.error(request, 'Please upload an Excel file only (.xlsx)')
            if request.user.groups.filter(name='TEACHER').exists():
                return redirect('teacher_page')
            else:
                return redirect('students')

        try:
            imported_data = dataset.load(
                new_student_file.read(), format='xlsx')
            successfully_imported = 0

            # Get the user's assigned classroom if they are in the TEACHER group
            if request.user.groups.filter(name='TEACHER').exists():
                if hasattr(request.user, 'teacher_profile'):
                    teacher = request.user.teacher_profile
                    teacher_classroom = Classroom.objects.filter(
                        teacher=teacher).first()
                    if not teacher_classroom:
                        messages.error(
                            request, 'You do not have an assigned classroom to import students into.')
                        return redirect('teachers_page')
                else:
                    messages.error(request, 'Teacher profile not found.')
                    return redirect('teachers_page')

            # Loop over each row in the dataset
            for i, data in enumerate(imported_data, start=1):
                try:
                    # Check and sanitize the LRN field
                    LRN_value = str(data[1]).strip()
                    if LRN_value is None or not LRN_value:
                        raise ValueError(
                            f"Row {i}: LRN is empty or None. Stopping further processing.")

                    # Assuming column index 13 has classroom data
                    classroom_identifier = data[13]
                    if classroom_identifier is None:
                        raise ValueError(
                            f"Row {i}: Classroom Identifier is None. Stopping further processing.")

                    # Get the classroom instance
                    try:
                        classroom_instance = Classroom.objects.get(
                            classroom=classroom_identifier)
                    except Classroom.DoesNotExist:
                        raise ValueError(f"Row {i}: Classroom with identifier {
                                         classroom_identifier} does not exist.")

                    # For teachers, ensure they only import students into their assigned classroom
                    if request.user.groups.filter(name='TEACHER').exists():
                        if classroom_instance != teacher_classroom:
                            raise ValueError(
                                f"Row {i}: You are only allowed to import students into your assigned classroom.")

                    # Handle birthday conversion
                    if isinstance(data[7], datetime.datetime):
                        converted_bday = data[7]
                    elif data[7] and isinstance(data[7], str):
                        try:
                            converted_bday = date_parser.parse(data[7])
                        except ValueError:
                            converted_bday = None
                    else:
                        converted_bday = None

                    # Construct the student object
                    student = Student(
                        LRN=int(data[1]),  # LRN as an integer
                        last_name=data[2],
                        first_name=data[3],
                        middle_name=data[4],
                        suffix_name=data[5],
                        status=data[6],
                        birthday=converted_bday,
                        religion=data[8],
                        other_religion=data[9],
                        strand=data[10],
                        age=data[11] if data[11] else None,
                        sem=data[12],
                        classroom=classroom_instance,
                        sex=data[15],
                        birth_place=data[16],
                        mother_tongue=data[17],
                        address=data[18],
                        father_name=data[19],
                        father_contact=data[20],
                        mother_name=data[21],
                        mother_contact=data[22],
                        guardian_name=data[23],
                        guardian_contact=data[24],
                        transfer_status=data[25],
                        household_income=data[26],
                        health_bmi=data[27] if data[27] else None,
                        general_average=data[28] if data[28] else None,
                        is_working_student=bool(data[29]),
                        is_returnee=bool(data[30]),
                        is_dropout=bool(data[31]),
                        is_4ps=bool(data[32]),
                        notes=data[33],
                        g7_school=data[36],
                        g7_schoolYear=data[37],
                        g7_section=data[38],
                        g7_general_average=data[39] if data[39] else None,
                        g7_adviser=data[40],
                        g7_adviserContact=data[41],
                        g8_school=data[43],
                        g8_schoolYear=data[44],
                        g8_section=data[45],
                        g8_general_average=data[46] if data[46] else None,
                        g8_adviser=data[47],
                        g8_adviserContact=data[48],
                        g9_school=data[50],
                        g9_schoolYear=data[51],
                        g9_section=data[52],
                        g9_general_average=data[53] if data[53] else None,
                        g9_adviser=data[54],
                        g9_adviserContact=data[55],
                        g10_school=data[57],
                        g10_schoolYear=data[58],
                        g10_section=data[59],
                        g10_general_average=data[60] if data[60] else None,
                        g10_adviser=data[61],
                        g10_adviserContact=data[62],
                        g11_school=data[64],
                        g11_schoolYear=data[65],
                        g11_section=data[66],
                        g11_general_average=data[67] if data[67] else None,
                        g11_adviser=data[68],
                        g11_adviserContact=data[69],
                        g12_school=data[71],
                        g12_schoolYear=data[72],
                        g12_section=data[73],
                        g12_general_average=data[74] if data[74] else None,
                        g12_adviser=data[75],
                        g12_adviserContact=data[76],
                    )

                    # Save the student to the database
                    student.save()
                    successfully_imported += 1
                    print(f"Row {i}: Successfully imported student:", student)

                except ValueError as ve:
                    error_message = f"Row {i}: {str(ve)}"
                    print(error_message)
                    messages.error(request, error_message)
                    break

                except IntegrityError as ie:
                    error_message = f"Row {i}: Integrity error - {str(ie)}"
                    print(error_message)
                    messages.error(request, error_message)
                    break

                except Exception as e:
                    error_message = f"Row {
                        i}: Error saving student data: {str(e)}"
                    print(error_message)
                    messages.error(request, error_message)
                    continue

            if successfully_imported > 0:
                messages.success(request, f"Successfully imported {
                                 successfully_imported} student(s) into the database.")

        except Exception as e:
            print(f"Error loading student/s from the file: {str(e)}")
            messages.error(
                request, f"Error loading students from the file: {str(e)}")

        # Redirect to different pages depending on user group
        if request.user.groups.filter(name='TEACHER').exists():
            return redirect('teacher_page')
        else:
            return redirect('students')

    return render(request, 'students')


def export_students_to_excel(request):
    try:
        # Fetch all student data
        students = Student.objects.all()

        # Load the template workbook from the new location
        template_path = os.path.join(
            'SIMS/static/media/VRCNHS_STUDENT_TEMPLATE.xlsx')
        existing_wb = load_workbook(template_path)
        sheet = existing_wb.active

        # Clear existing data in the template
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

        start_row = 2
        start_column = 2

        # NamedStyle for date formatting
        date_style = NamedStyle(name='date_style', number_format='MM-DD-YYYY')

        # Define the list representing the order of columns in the Excel file
        excel_columns_order = [
            'LRN', 'last_name', 'first_name', 'middle_name', 'suffix_name', 'status', 'birthday',
            'religion', 'other_religion', 'strand', 'age', 'sem', 'classroom', 'sex',
            'birth_place', 'mother_tongue', 'address', 'father_name', 'father_contact', 'mother_name',
            'mother_contact', 'guardian_name', 'guardian_contact', 'transfer_status', 'household_income',
            'is_returnee', 'is_dropout', 'is_working_student', 'health_bmi', 'general_average',
            'is_4ps', 'notes',
            '',  # Blank column
            '',
            # Grade 7
            'g7_school', 'g7_schoolYear', 'g7_section', 'g7_general_average', 'g7_adviser', 'g7_adviserContact',
            '',  # Blank column
            # Grade 8
            'g8_school', 'g8_schoolYear', 'g8_section', 'g8_general_average', 'g8_adviser', 'g8_adviserContact',
            '',  # Blank column
            # Grade 9
            'g9_school', 'g9_schoolYear', 'g9_section', 'g9_general_average', 'g9_adviser', 'g9_adviserContact',
            '',  # Blank column
            # Grade 10
            'g10_school', 'g10_schoolYear', 'g10_section', 'g10_general_average', 'g10_adviser', 'g10_adviserContact',
            '',  # Blank column
            # Grade 11
            'g11_school', 'g11_schoolYear', 'g11_section', 'g11_general_average', 'g11_adviser', 'g11_adviserContact',
            '',  # Blank column
            # Grade 12
            'g12_school', 'g12_schoolYear', 'g12_section', 'g12_general_average', 'g12_adviser', 'g12_adviserContact',
        ]

        # Loop through each student and write the data to the template
        for row_num, student in enumerate(students, start_row):
            for col_num, attribute in enumerate(excel_columns_order, start_column):
                col_letter = get_column_letter(col_num)

                if attribute == '':
                    # Skip blank columns
                    continue

                field_value = getattr(student, attribute, None)

                if attribute in ['health_bmi', 'general_average'] and field_value is not None:
                    # Ensure health_bmi and general_average are converted to float if they exist
                    field_value = float(field_value)

                elif attribute == 'birthday' and field_value is not None:
                    # Export date in MM-DD-YYYY format
                    field_value = student.birthday.strftime('%m-%d-%Y')
                    sheet[f"{col_letter}{row_num}"] = field_value
                    sheet[f"{col_letter}{row_num}"].number_format = 'MM-DD-YYYY'
                    continue

                elif isinstance(field_value, bool):
                    # Convert boolean values to Yes/No
                    field_value = 'Yes' if field_value else 'No'

                elif field_value is not None:
                    field_value = str(field_value)

                else:
                    field_value = ""

                sheet[f"{col_letter}{row_num}"] = field_value

        # Create an HTTP response to allow the user to download the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students_data_updated.xlsx'
        existing_wb.save(response)

        messages.success(request, "Data successfully exported to Excel!")
        return response

    except Exception as e:
        # Handle any exceptions that may occur during export
        print(f"Error exporting data to Excel: {str(e)}")
        messages.error(
            request, "An error occurred while exporting data to Excel. Please try again.")

    return redirect('students')


def export_classroom_students_to_excel(request):
    try:
        # Ensure the user is under the TEACHER group
        if request.user.groups.filter(name='TEACHER').exists():
            # Get the teacher instance and assigned classroom
            teacher = get_object_or_404(Teacher, user=request.user)
            classroom = get_object_or_404(Classroom, teacher=teacher)

            # Get students belonging to the classroom
            students = Student.objects.filter(classroom=classroom)

            # Load the Excel template workbook
            template_path = 'SIMS/static/media/VRCNHS_STUDENT_TEMPLATE.xlsx'
            existing_wb = load_workbook(template_path)
            sheet = existing_wb.active

            # Clear existing data from the sheet
            for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
                for cell in row:
                    cell.value = None

            # Start row and column positions for populating student data
            start_row = 2
            start_column = 2

            # Define the order of columns in the Excel file as per the Student model
            excel_columns_order = [
                'LRN', 'last_name', 'first_name', 'middle_name', 'suffix_name', 'status', 'birthday',
                'religion', 'other_religion', 'strand', 'age', 'sem', 'classroom', 'sex', 'birth_place',
                'mother_tongue', 'address', 'father_name', 'father_contact', 'mother_name', 'mother_contact',
                'guardian_name', 'guardian_contact', 'transfer_status', 'household_income', 'health_bmi',
                'general_average', 'is_working_student', 'is_returnee', 'is_dropout', 'is_4ps', 'notes',
                '',  # Blank column
                '',
                # Grade 7
                'g7_school', 'g7_schoolYear', 'g7_section', 'g7_general_average', 'g7_adviser', 'g7_adviserContact',
                '',  # Blank column
                # Grade 8
                'g8_school', 'g8_schoolYear', 'g8_section', 'g8_general_average', 'g8_adviser', 'g8_adviserContact',
                '',  # Blank column
                # Grade 9
                'g9_school', 'g9_schoolYear', 'g9_section', 'g9_general_average', 'g9_adviser', 'g9_adviserContact',
                '',  # Blank column
                # Grade 10
                'g10_school', 'g10_schoolYear', 'g10_section', 'g10_general_average', 'g10_adviser', 'g10_adviserContact',
                '',  # Blank column
                # Grade 11
                'g11_school', 'g11_schoolYear', 'g11_section', 'g11_general_average', 'g11_adviser', 'g11_adviserContact',
                '',  # Blank column
                # Grade 12
                'g12_school', 'g12_schoolYear', 'g12_section', 'g12_general_average', 'g12_adviser', 'g12_adviserContact',
            ]

            # Write student data to the Excel sheet
            for row_num, student in enumerate(students, start_row):
                for col_num, attribute in enumerate(excel_columns_order, start_column):
                    col_letter = get_column_letter(col_num)

                    if attribute == '':
                        # Skip blank columns
                        continue

                    # Extract field value based on attribute
                    field_value = getattr(student, attribute, None)

                    # Handle specific attributes like classroom or gradelevel
                    if attribute == 'classroom':
                        field_value = student.classroom.classroom if student.classroom else ''
                    elif attribute == 'birthday' and field_value:
                        # Format birthday in MM-DD-YYYY format
                        field_value = field_value.strftime('%m-%d-%Y')

                    # Set value in the cell
                    sheet[f"{col_letter}{row_num}"] = field_value

            # Create HTTP response with the generated Excel file
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=classroom_students_data.xlsx'
            existing_wb.save(response)

            messages.success(
                request, "Classroom data successfully exported to Excel!")

            return response

        else:
            messages.error(
                request, "You are not authorized to export classroom data.")
            return redirect('teacher_page')

    except Exception as e:
        print(f"Error exporting classroom data to Excel: {str(e)}")
        messages.error(
            request, "An error occurred while exporting classroom data to Excel. Please try again.")

    return redirect('teacher_page')


def export_and_delete_students_for_departure(request):
    try:
        # Retrieve students for departure
        students_for_departure = Student.objects.filter(
            Q(status='For Graduation') | Q(
                status='For Dropout') | Q(status='For Transfer'),
        )

        # Load the template workbook
        template_path = 'SIMS/static/media/VRCNHS_STUDENT_TEMPLATE.xlsx'
        existing_wb = load_workbook(template_path)
        sheet = existing_wb.active

        # Clear existing data
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

        start_row = 2
        start_column = 2

        date_style = NamedStyle(name='date_style', number_format='MM-DD-YYYY')

        # Define a list representing the order of columns in the Excel file
        excel_columns_order = [
            'LRN', 'last_name', 'first_name', 'middle_name', 'suffix_name', 'status', 'birthday',
            'religion', 'other_religion', 'strand', 'age', 'sem', 'classroom', 'sex',
            'birth_place', 'mother_tongue', 'address', 'father_name', 'father_contact', 'mother_name',
            'mother_contact', 'guardian_name', 'guardian_contact', 'transfer_status', 'household_income',
            'is_returnee', 'is_dropout', 'is_working_student', 'health_bmi', 'general_average',
            'is_4ps', 'notes',
            '',  # Blank column
            '',
            # Grade 7
            'g7_school', 'g7_schoolYear', 'g7_section', 'g7_general_average', 'g7_adviser', 'g7_adviserContact',
            '',  # Blank column
            # Grade 8
            'g8_school', 'g8_schoolYear', 'g8_section', 'g8_general_average', 'g8_adviser', 'g8_adviserContact',
            '',  # Blank column
            # Grade 9
            'g9_school', 'g9_schoolYear', 'g9_section', 'g9_general_average', 'g9_adviser', 'g9_adviserContact',
            '',  # Blank column
            # Grade 10
            'g10_school', 'g10_schoolYear', 'g10_section', 'g10_general_average', 'g10_adviser', 'g10_adviserContact',
            '',  # Blank column
            # Grade 11
            'g11_school', 'g11_schoolYear', 'g11_section', 'g11_general_average', 'g11_adviser', 'g11_adviserContact',
            '',  # Blank column
            # Grade 12
            'g12_school', 'g12_schoolYear', 'g12_section', 'g12_general_average', 'g12_adviser', 'g12_adviserContact',
        ]

        for row_num, student in enumerate(students_for_departure, start_row):
            for col_num, attribute in enumerate(excel_columns_order, start_column):
                col_letter = get_column_letter(col_num)

                if attribute == '':
                    # Skip blank columns
                    continue

                field_value = getattr(student, attribute, None)

                if attribute in ['LRN',]:
                    if field_value is not None:
                        field_value = int(field_value)
                elif attribute in ['health_bmi', 'general_average']:
                    if field_value is not None:
                        field_value = float(field_value)
                elif attribute == 'birthday':
                    # Export date in MM-DD-YYYY format
                    field_value = student.birthday.strftime(
                        '%m-%d-%Y') if student.birthday else None
                    sheet[f"{col_letter}{row_num}"] = field_value
                    sheet[f"{col_letter}{row_num}"].number_format = 'MM-DD-YYYY'
                else:
                    if field_value is not None:
                        field_value = str(field_value)
                    else:
                        field_value = ""

                sheet[f"{col_letter}{row_num}"] = field_value

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students_for_departure.xlsx'
        existing_wb.save(response)

        # Delete students after successful export
        students_for_departure.delete()

        messages.success(
            request, "Data successfully exported to Excel and students deleted!")

        return response

    except Exception as e:
        print(f"Error exporting data to Excel: {str(e)}")
        messages.error(
            request, "An error occurred while exporting data to Excel. Please try again.")
        # Redirect to sectioning page in case of an error
        return redirect('sectioning')
